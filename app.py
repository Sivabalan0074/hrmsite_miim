# =============================================================================
#  MIIM HR System — Flask Backend
#  app.py  |  Full integrated version
#
#  Features:
#    ✅ SQLite database (miim.db)
#    ✅ Excel builder  — Sheet 1: All Employees | Sheet 2+: Per Department
#    ✅ Excel → DB sync on file save (auto-watcher every 10 s)
#    ✅ Manual sync via POST /sync_from_excel
#    ✅ REST API: login, employees CRUD, approve, reset password, photo upload
#    ✅ Serves 1_miim.html (home) and 2_hr_dashboard.html (HR panel)
# =============================================================================

from flask import Flask, request, jsonify, send_from_directory, send_file
import sqlite3, hashlib, os, re, threading, time, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText as _MIMEText
from datetime import datetime
from collections import defaultdict

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024   # 16 MB upload limit

# =============================================================================
#  EMAIL CONFIG — Hostinger SMTP
# =============================================================================
SMTP_HOST  = "smtp.hostinger.com"
SMTP_PORT  = 587
SMTP_USER  = "chethan.r@miim.co.in"
SMTP_PASS  = "Miim@123#"
SMTP_FROM  = "MIIM HR <chethan.r@miim.co.in>"
PORTAL_URL = "http://localhost:5005"


def send_email(to_addr: str, subject: str, html_body: str) -> tuple:
    """Send an HTML email via Hostinger SMTP. Returns (ok: bool, error: str)."""
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = SMTP_FROM
        msg["To"]      = to_addr
        msg.attach(_MIMEText(html_body, "html"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as s:
            s.ehlo(); s.starttls(); s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_USER, to_addr, msg.as_string())
        return True, ""
    except Exception as ex:
        return False, str(ex)


def reset_email_html(name: str, reset_link: str) -> str:
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0a0a0a;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#0a0a0a;padding:40px 0;">
    <tr><td align="center">
      <table width="520" cellpadding="0" cellspacing="0"
             style="background:#111;border:1px solid rgba(255,107,0,0.3);border-radius:4px;">
        <tr><td style="background:#1a1a1a;padding:28px 36px;border-bottom:2px solid #FF6B00;">
          <p style="margin:0;font-size:22px;font-weight:700;color:#FF6B00;letter-spacing:3px;text-transform:uppercase;">MIIM</p>
          <p style="margin:4px 0 0;font-size:12px;color:#888;letter-spacing:2px;text-transform:uppercase;">Mission Impossible Industrial Management</p>
        </td></tr>
        <tr><td style="padding:36px;">
          <p style="margin:0 0 18px;font-size:16px;color:#ccc;">Hello, <strong style="color:#fff;">{name}</strong></p>
          <p style="margin:0 0 24px;font-size:14px;color:#aaa;line-height:1.7;">Your HR team has initiated a password reset for your MIIM Employee Portal account. Click the button below to set your new password.</p>
          <table cellpadding="0" cellspacing="0" style="margin:0 auto 28px;">
            <tr><td style="background:#FF6B00;">
              <a href="{reset_link}" style="display:block;padding:14px 36px;color:#000;font-size:14px;font-weight:700;text-decoration:none;letter-spacing:2px;text-transform:uppercase;">Reset My Password</a>
            </td></tr>
          </table>
          <p style="margin:0 0 8px;font-size:12px;color:#666;">Or copy this link into your browser:</p>
          <p style="margin:0 0 28px;word-break:break-all;"><a href="{reset_link}" style="font-size:12px;color:#FF6B00;">{reset_link}</a></p>
          <p style="margin:0;font-size:12px;color:#555;line-height:1.6;">If you did not request this, please ignore this email or contact HR immediately.</p>
        </td></tr>
        <tr><td style="background:#0a0a0a;padding:18px 36px;border-top:1px solid #222;">
          <p style="margin:0;font-size:11px;color:#444;">&copy; MIIM — Mission Impossible Industrial Management &nbsp;|&nbsp; Automated message.</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

# =============================================================================
#  CONFIGURATION — edit paths here if needed
# =============================================================================
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_FILE     = os.path.join(BASE_DIR, "miim.db")

# Storage folder sits next to app.py — works on any machine / any user
STORAGE_DIR = os.path.join(BASE_DIR, "miim_storage")
PHOTOS_DIR  = os.path.join(STORAGE_DIR, "employee_photos")
EXCEL_FILE  = os.path.join(STORAGE_DIR, "MIIM_Employees.xlsx")

os.makedirs(PHOTOS_DIR, exist_ok=True)

# =============================================================================
#  EXCEL COLUMN DEFINITION
#  Used by both the builder (DB → Excel) and the sync reader (Excel → DB).
#  If you add/move a column, update EXCEL_HEADERS, EXCEL_COL_WIDTHS,
#  and EXCEL_COL_MAP together.
# =============================================================================
EXCEL_HEADERS = [
    "ID", "USERNAME", "FULL NAME", "ROLE / DESIGNATION",
    "DEPARTMENT", "MANAGER", "EMAIL", "PHONE",
    "EMP TYPE", "STATUS", "JOIN DATE", "ADDRESS"
]
EXCEL_COL_WIDTHS = [8, 18, 22, 28, 18, 18, 28, 14, 14, 10, 14, 30]

# 1-based column index → DB field name
EXCEL_COL_MAP = {
    1:  "id",
    2:  "username",
    3:  "name",
    4:  "role",
    5:  "department",
    6:  "manager",
    7:  "email",
    8:  "phone",
    9:  "emp_type",
    10: "status",
    11: "joindate",
    12: "address",
}

# Fields the Excel sync is allowed to write back to the DB
# (id and username are read-only from Excel)
EXCEL_UPDATABLE = {
    "name", "role", "department", "manager",
    "email", "phone", "emp_type", "status", "joindate", "address"
}


# =============================================================================
#  HELPERS
# =============================================================================
def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def safe_sheet_name(name: str) -> str:
    """Strip illegal Excel sheet-name chars and truncate to 31."""
    return re.sub(r'[\\/*?:\[\]]', '_', str(name))[:31]


def fetch_all_employees() -> list:
    """Return every user row as a list of dicts, ordered by dept then id."""
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, username, name, role, department, email,
               phone, address, access, status, created_at,
               approval_status, manager, empid, joindate, emp_type
        FROM users
        ORDER BY department, id
    """)
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return rows


# =============================================================================
#  DATABASE INITIALISATION
# =============================================================================
def init_db():
    conn   = get_db()
    cursor = conn.cursor()

    # Create table (safe if already exists)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id              INTEGER  PRIMARY KEY AUTOINCREMENT,
            username        TEXT     NOT NULL UNIQUE,
            password        TEXT     NOT NULL,
            name            TEXT,
            role            TEXT,
            department      TEXT,
            email           TEXT,
            phone           TEXT,
            address         TEXT,
            photo           TEXT,
            access          TEXT     DEFAULT 'Employee',
            status          TEXT     DEFAULT 'Active',
            approval_status TEXT     DEFAULT 'approved',
            manager         TEXT     DEFAULT '',
            empid           TEXT     DEFAULT '',
            joindate        TEXT     DEFAULT '',
            emp_type        TEXT     DEFAULT 'Regular',
            created_at      DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Migrate older DBs that may be missing newer columns
    migration_cols = [
        ("approval_status",         "'approved'"),
        ("manager",                  "''"),
        ("empid",                    "''"),
        ("joindate",                 "''"),
        ("emp_type",                 "'Regular'"),
        ("password_change_required", "0"),
        ("reset_token",              "NULL"),
        ("reset_token_expiry",       "NULL"),
    ]
    for col, default in migration_cols:
        try:
            cursor.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT DEFAULT {default}")
        except Exception:
            pass   # column already exists

    # Seed default admin account ONLY if it does not already exist.
    # First-time: password = "admin", forced to change on first login.
    # Once the admin has changed their password the row already exists
    # (INSERT OR IGNORE is a no-op) so the custom password is preserved.
    cursor.execute("""
        INSERT OR IGNORE INTO users
            (username, password, name, role, department, email,
             access, approval_status, status, password_change_required)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        "admin", hash_password("admin"),
        "Admin", "Administrator",
        "HR Department", "admin@miim.in",
        "Full Access", "approved", "Active", 1
    ))

    # Only fix the approval/status flags — never touch the password here.
    # This lets the admin keep the password they set on first login.
    cursor.execute("""
        UPDATE users SET
            approval_status = 'approved',
            status = 'Active'
        WHERE username = 'admin'
          AND (approval_status != 'approved' OR status != 'Active')
    """)

    conn.commit()
    conn.close()

    print("✅  MIIM Database ready.")
    print(f"    DB      : {DB_FILE}")
    print(f"    Storage : {STORAGE_DIR}")
    print(f"    Photos  : {PHOTOS_DIR}")
    print(f"    Excel   : {EXCEL_FILE}")

    # Start background Excel-file watcher
    t = threading.Thread(target=_excel_watcher, daemon=True)
    t.start()
    print("✅  Excel file-watcher started (checks every 10 s).")


# =============================================================================
#  EXCEL BUILDER  (DB → Excel)
#
#  Sheet layout (rebuilt from scratch on every call):
#    [0] "All Employees"  — master list of every employee
#    [1] "<Dept A>"       — employees in dept A  (alphabetical)
#    [2] "<Dept B>"       — employees in dept B
#    ...
# =============================================================================
def build_excel():
    """
    Rebuild MIIM_Employees.xlsx from the current DB state.
    Returns (True, excel_path) on success, (False, error_str) on failure.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # ── Shared style objects ──────────────────────────────────────────
        DARK   = "1A1D23"
        ORANGE = "E8721C"

        T_FONT  = Font(name="Arial", bold=True,  color="FFFFFF", size=13)
        H_FONT  = Font(name="Arial", bold=True,  color="FFFFFF", size=10)
        D_FONT  = Font(name="Arial",              color="374151", size=10)
        B_FONT  = Font(name="Arial", bold=True,  color="1A1D23", size=10)
        M_FONT  = Font(name="Arial", italic=True, color="6B7280", size=9)

        T_FILL  = PatternFill("solid", fgColor=DARK)
        H_FILL  = PatternFill("solid", fgColor="374151")
        ODD     = PatternFill("solid", fgColor="FFFFFF")
        EVEN    = PatternFill("solid", fgColor="F3F4F6")
        O_FILL  = PatternFill("solid", fgColor=ORANGE)

        _side   = Side(style="thin", color="E5E7EB")
        BDR     = Border(left=_side, right=_side, top=_side, bottom=_side)
        CTR     = Alignment(horizontal="center", vertical="center")
        LFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        # ── Helper: write title + subtitle rows ──────────────────────────
        def _title(ws, title, n, subtitle=""):
            last = get_column_letter(n)
            ws.merge_cells(f"A1:{last}1")
            ws["A1"] = title
            ws["A1"].font = T_FONT; ws["A1"].fill = T_FILL; ws["A1"].alignment = CTR
            ws.row_dimensions[1].height = 34

            ws.merge_cells(f"A2:{last}2")
            ws["A2"] = subtitle
            ws["A2"].font = M_FONT; ws["A2"].alignment = CTR
            ws.row_dimensions[2].height = 16

        # ── Helper: write column-header row (row 3) ──────────────────────
        def _headers(ws, headers, widths):
            for c, (h, w) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=3, column=c, value=h)
                cell.font = H_FONT; cell.fill = H_FILL
                cell.alignment = CTR; cell.border = BDR
                ws.column_dimensions[get_column_letter(c)].width = w
            ws.row_dimensions[3].height = 22

        # ── Helper: extract ordered values from an employee dict ─────────
        def _vals(emp):
            return [
                emp.get("id",          ""),
                emp.get("username",    ""),
                emp.get("name",        "") or emp.get("username", ""),
                emp.get("role",        ""),
                emp.get("department",  ""),
                emp.get("manager",     ""),
                emp.get("email",       ""),
                emp.get("phone",       ""),
                emp.get("emp_type",    "Regular"),
                emp.get("status",      "Active"),
                emp.get("joindate",    "") or (emp.get("created_at", "")[:10] if emp.get("created_at") else ""),
                emp.get("address",     ""),
            ]

        # ── Helper: write data rows starting at start_row ────────────────
        def _data(ws, emps, start_row=4):
            for ri, emp in enumerate(emps, start_row):
                fill = ODD if ri % 2 else EVEN
                for ci, v in enumerate(_vals(emp), 1):
                    cell = ws.cell(row=ri, column=ci, value=(v if v != "" else ""))
                    cell.fill = fill; cell.border = BDR
                    if   ci == 3:                      # Full Name — bold
                        cell.font = B_FONT; cell.alignment = LFT
                    elif ci == 10:                     # Status — green / red
                        cell.font = Font(name="Arial", bold=True, size=10,
                                         color="15803D" if str(v) == "Active" else "DC2626")
                        cell.alignment = CTR
                    elif ci in (4, 5, 6, 7, 12):      # text fields — left
                        cell.font = D_FONT; cell.alignment = LFT
                    else:
                        cell.font = D_FONT; cell.alignment = CTR
                ws.row_dimensions[ri].height = 20

        # ── Helper: orange summary row at bottom ─────────────────────────
        def _summary(ws, emps, start_row, n):
            sr   = start_row + len(emps)
            last = get_column_letter(n)
            ws.merge_cells(f"A{sr}:{last}{sr}")
            active = sum(1 for e in emps if (e.get("status") or "Active") == "Active")
            ws[f"A{sr}"] = (
                f"Total: {len(emps)}   |   Active: {active}"
                f"   |   Inactive: {len(emps)-active}"
                f"   |   Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            ws[f"A{sr}"].font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            ws[f"A{sr}"].fill      = O_FILL
            ws[f"A{sr}"].alignment = CTR
            ws.row_dimensions[sr].height = 18

        # ── Fetch all employees from DB ───────────────────────────────────
        all_emps = fetch_all_employees()
        n        = len(EXCEL_HEADERS)
        ts       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # ════════════════════════════════════════════════════════════════
        #  SHEET 1 — "All Employees"  (always position 0)
        # ════════════════════════════════════════════════════════════════
        ws_all = wb.create_sheet("All Employees", 0)
        _title(ws_all, "MIIM – Employee Master List", n,
               f"Generated: {ts}   |   Total employees: {len(all_emps)}")
        _headers(ws_all, EXCEL_HEADERS, EXCEL_COL_WIDTHS)
        _data(ws_all, all_emps)
        _summary(ws_all, all_emps, 4, n)
        ws_all.freeze_panes = "A4"

        # ════════════════════════════════════════════════════════════════
        #  SHEET 2+  — one sheet per department (sorted A-Z)
        # ════════════════════════════════════════════════════════════════
        dept_map = defaultdict(list)
        for emp in all_emps:
            dept = (emp.get("department") or "Unknown").strip()
            dept_map[dept].append(emp)

        for dept in sorted(dept_map.keys()):
            emps  = dept_map[dept]
            ws_d  = wb.create_sheet(safe_sheet_name(dept))
            _title(ws_d, f"{dept} Department", n,
                   f"Generated: {ts}   |   Employees: {len(emps)}")
            _headers(ws_d, EXCEL_HEADERS, EXCEL_COL_WIDTHS)
            _data(ws_d, emps)
            _summary(ws_d, emps, 4, n)
            ws_d.freeze_panes = "A4"

        wb.save(EXCEL_FILE)
        return True, EXCEL_FILE

    except Exception as ex:
        return False, str(ex)


# =============================================================================
#  EXCEL → DB SYNC
#
#  Reads "All Employees" sheet (index 0), row 4 onwards.
#  • Row has an ID   → update changed fields in DB
#  • Row has no ID   → insert as new employee (username + name required)
#  Deletions in Excel are IGNORED — delete employees via the HR dashboard.
# =============================================================================
def sync_excel_to_db():
    """
    Push edits made in the Excel file back into the database.
    Returns (changes: int, log: list[str]).
    """
    if not os.path.exists(EXCEL_FILE):
        return 0, ["Excel file not found."]

    try:
        from openpyxl import load_workbook
        wb  = load_workbook(EXCEL_FILE, data_only=True)
        ws  = wb.worksheets[0]      # "All Employees" is always index 0

        conn    = get_db()
        cursor  = conn.cursor()
        changes = 0
        log     = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not any(row):
                continue    # completely blank row — skip

            row_id   = row[0]
            username = str(row[1] or "").strip()
            name     = str(row[2] or "").strip()

            if not row_id and not username:
                continue

            # Build field dict from Excel row
            excel_data = {}
            for col_idx, field in EXCEL_COL_MAP.items():
                if field in EXCEL_UPDATABLE:
                    val = row[col_idx - 1]   # row is 0-based
                    excel_data[field] = str(val).strip() if val is not None else ""

            # ── UPDATE existing employee ─────────────────────────────────
            if row_id:
                cursor.execute("SELECT * FROM users WHERE id = ?", (int(row_id),))
                db_row = cursor.fetchone()
                if not db_row:
                    log.append(f"ID {row_id}: not found in DB — skipped.")
                    continue

                diff = {
                    f: v for f, v in excel_data.items()
                    if str(db_row[f] or "").strip() != v
                }
                if diff:
                    set_clause = ", ".join(f"{k} = ?" for k in diff)
                    cursor.execute(
                        f"UPDATE users SET {set_clause} WHERE id = ?",
                        list(diff.values()) + [int(row_id)]
                    )
                    changes += 1
                    log.append(f"Updated ID {row_id} ({db_row['username']}): {list(diff.keys())}")

            # ── INSERT new employee added directly in Excel ───────────────
            else:
                if not username or not name:
                    log.append("New row skipped — username and name are both required.")
                    continue
                default_pw = hash_password(username)   # password = username, must change on login
                try:
                    cursor.execute("""
                        INSERT INTO users
                          (username, password, name, role, department, email,
                           phone, address, manager, joindate, emp_type,
                           status, approval_status, access, password_change_required)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        username, default_pw, name,
                        excel_data.get("role",       ""),
                        excel_data.get("department", ""),
                        excel_data.get("email",      ""),
                        excel_data.get("phone",      ""),
                        excel_data.get("address",    ""),
                        excel_data.get("manager",    ""),
                        excel_data.get("joindate",   ""),
                        excel_data.get("emp_type",   "Regular"),
                        excel_data.get("status",     "Active"),
                        "approved", "Employee",
                        1,   # force password change on first login
                    ))
                    changes += 1
                    log.append(f"Inserted from Excel: {username}")
                except sqlite3.IntegrityError:
                    log.append(f"Username '{username}' already exists — skipped.")

        conn.commit()
        conn.close()

        if changes:
            build_excel()   # rebuild so dept sheets stay in sync

        return changes, log

    except Exception as ex:
        return 0, [f"Sync error: {ex}"]


# =============================================================================
#  EXCEL FILE WATCHER  (background daemon thread)
#  Polls EXCEL_FILE every 10 seconds; if mtime changes → sync to DB.
# =============================================================================
_last_excel_mtime = None

def _excel_watcher():
    global _last_excel_mtime
    while True:
        time.sleep(10)
        try:
            if not os.path.exists(EXCEL_FILE):
                continue
            mtime = os.path.getmtime(EXCEL_FILE)
            if _last_excel_mtime is None:
                _last_excel_mtime = mtime
                continue
            if mtime != _last_excel_mtime:
                _last_excel_mtime = mtime
                changes, log = sync_excel_to_db()
                if changes:
                    print(f"\n[Excel Watcher] ✅ {changes} change(s) synced to DB:")
                    for line in log:
                        print(f"  • {line}")
                else:
                    print("[Excel Watcher] File changed — no DB updates needed.")
        except Exception as e:
            print(f"[Excel Watcher] Error: {e}")


# =============================================================================
#  ROUTES — Static pages
# =============================================================================
@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "1_miim.html")

@app.route("/hr")
def hr_dashboard():
    return send_from_directory(BASE_DIR, "3_hr_dashboard.html")

@app.route("/set-password")
def set_password():
    return send_from_directory(BASE_DIR, "2_hr_dashboard.html")

@app.route("/<path:filename>")
def serve_static(filename):
    allowed_ext = (".html", ".css", ".js", ".png", ".jpg", ".jpeg", ".ico", ".svg", ".woff", ".woff2")
    if filename.endswith(allowed_ext):
        return send_from_directory(BASE_DIR, filename)
    return jsonify({"error": "Not found"}), 404


# =============================================================================
#  ROUTES — Auth
# =============================================================================
@app.route("/login", methods=["POST"])
def login():
    data     = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"success": False, "message": "Username and password required."}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, username, name, role, department,
               email, phone, photo, access, status,
               approval_status, password_change_required
        FROM users
        WHERE username = ? AND password = ?
    """, (username, hash_password(password)))
    user = cursor.fetchone()
    conn.close()

    if not user:
        return jsonify({"success": False, "message": "Invalid username or password."}), 401

    if (user["status"] or "Active") != "Active":
        return jsonify({"success": False, "message": "Your account is inactive. Please contact HR."}), 403

    if (user["approval_status"] or "approved") not in ("approved", ""):
        return jsonify({"success": False, "message": "Your account is pending approval."}), 403

    photo_url = user["photo"]
    if photo_url and not photo_url.startswith("data:"):
        photo_url = f"/photo/{user['id']}"

    must_change = bool(user["password_change_required"])

    return jsonify({
        "success": True,
        "must_change_password": must_change,
        "message": f"Welcome, {user['name']}!",
        "user": {
            "id":         user["id"],
            "username":   user["username"],
            "name":       user["name"],
            "role":       user["role"],
            "department": user["department"],
            "email":      user["email"],
            "phone":      user["phone"],
            "photo":      photo_url,
            "access":     user["access"],
        }
    })


# =============================================================================
#  ROUTES — Employees (CRUD)
# =============================================================================

# ── GET ALL ──────────────────────────────────────────────────────────────────
@app.route("/employees", methods=["GET"])
def get_employees():
    rows = fetch_all_employees()
    for emp in rows:
        if emp.get("photo") and not emp["photo"].startswith("data:"):
            emp["photo"] = f"/photo/{emp['id']}"
    return jsonify({"success": True, "employees": rows})


# ── ADD ───────────────────────────────────────────────────────────────────────
@app.route("/add_employee", methods=["POST"])
def add_employee():
    data = request.get_json() or {}
    for required in ("username", "name"):
        if not data.get(required):
            return jsonify({"success": False, "message": f"'{required}' is required."}), 400

    username = data["username"]
    # Default password = username (employee must change on first login)
    raw_password = data.get("password") or username

    conn   = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO users
              (username, password, name, role, department, email,
               phone, address, photo, access,
               approval_status, manager, empid, joindate, emp_type,
               password_change_required)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            username,
            hash_password(raw_password),
            data["name"],
            data.get("role",            data.get("desig", "Employee")),
            data.get("department",      ""),
            data.get("email",           ""),
            data.get("phone",           ""),
            data.get("address",         ""),
            data.get("photo",           ""),
            data.get("access",          "Employee"),
            data.get("approval_status", "pending"),
            data.get("manager",         ""),
            data.get("empid",           ""),
            data.get("joindate",        ""),
            data.get("emp_type",        "Regular"),
            1,   # force password change on first login
        ))
        conn.commit()
        new_id = cursor.lastrowid
        conn.close()
        build_excel()
        return jsonify({
            "success": True,
            "message": (
                f"Employee '{data['name']}' added. "
                f"Default password is their username ('{username}'). "
                "They will be required to set a new password on first login."
            ),
            "id": new_id
        })
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"success": False, "message": "Username already exists."}), 409


# ── UPDATE ────────────────────────────────────────────────────────────────────
@app.route("/update_employee/<int:user_id>", methods=["PUT"])
def update_employee(user_id):
    data   = request.get_json() or {}
    conn   = get_db()
    cursor = conn.cursor()

    allowed = [
        "name", "role", "department", "email", "phone", "address",
        "access", "status", "approval_status",
        "manager", "empid", "joindate", "emp_type"
    ]
    updates = {k: data[k] for k in allowed if k in data}
    if "password" in data and data["password"]:
        updates["password"] = hash_password(data["password"])

    if not updates:
        conn.close()
        return jsonify({"success": False, "message": "No fields to update."}), 400

    set_clause = ", ".join(f"{k} = ?" for k in updates)
    cursor.execute(
        f"UPDATE users SET {set_clause} WHERE id = ?",
        list(updates.values()) + [user_id]
    )
    conn.commit()
    conn.close()
    build_excel()
    return jsonify({"success": True, "message": "Employee updated successfully."})


# ── DELETE ────────────────────────────────────────────────────────────────────
@app.route("/delete_employee/<int:user_id>", methods=["DELETE"])
def delete_employee(user_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username, photo FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()

    # Remove photo file if it exists
    if row and row["photo"] and not row["photo"].startswith("data:"):
        fpath = os.path.join(PHOTOS_DIR, row["photo"])
        if os.path.exists(fpath):
            try: os.remove(fpath)
            except: pass

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    build_excel()
    return jsonify({"success": True, "message": "Employee deleted."})


# ── APPROVE ───────────────────────────────────────────────────────────────────
@app.route("/approve_employee/<int:user_id>", methods=["POST"])
def approve_employee(user_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET approval_status = 'approved' WHERE id = ?", (user_id,)
    )
    conn.commit()
    conn.close()
    build_excel()
    return jsonify({"success": True, "message": "Employee approved."})


# ── RESET PASSWORD (by HR — sets temp pw, marks must-change) ─────────────────
@app.route("/reset_password", methods=["POST"])
def reset_password():
    data     = request.get_json() or {}
    user_id  = data.get("empId")
    password = data.get("password", "")
    if not user_id or not password:
        return jsonify({"success": False, "message": "empId and password are required."}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET password = ?, password_change_required = 1 WHERE id = ?",
        (hash_password(password), user_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Password reset. User must change on next login."})


# ── CHANGE PASSWORD (by the employee themselves) ──────────────────────────────
@app.route("/change_password", methods=["POST"])
def change_password():
    """
    Employee sets their new permanent password.
    Body: { username, current_password, new_password }
    Clears password_change_required flag on success.
    """
    data         = request.get_json() or {}
    username     = data.get("username", "").strip()
    current_pw   = data.get("current_password", "")
    new_pw       = data.get("new_password", "")

    if not username or not current_pw or not new_pw:
        return jsonify({"success": False, "message": "username, current_password and new_password required."}), 400

    if len(new_pw) < 6:
        return jsonify({"success": False, "message": "New password must be at least 6 characters."}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM users WHERE username = ? AND password = ? AND status = 'Active'",
        (username, hash_password(current_pw))
    )
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "Current password is incorrect."}), 401

    cursor.execute(
        "UPDATE users SET password = ?, password_change_required = 0, reset_token = NULL, reset_token_expiry = NULL WHERE id = ?",
        (hash_password(new_pw), row["id"])
    )
    conn.commit()
    conn.close()
    build_excel()   # keep Excel in sync (no sensitive data stored there)
    return jsonify({"success": True, "message": "Password changed successfully."})


# ── FORGOT PASSWORD — employee requests reset via HR ─────────────────────────
@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    """
    Employee submits username to request a password reset.
    - Username NOT found  → not_registered: True  (UI shows popup)
    - Username found      → generates token, sends email, returns success
    Body: { username }
    """
    import secrets, string
    data     = request.get_json() or {}
    username = data.get("username", "").strip()
    if not username:
        return jsonify({"success": False, "message": "Username required."}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, email, name FROM users WHERE username = ? AND status = 'Active'", (username,))
    row = cursor.fetchone()

    if not row:
        conn.close()
        return jsonify({
            "success":        False,
            "not_registered": True,
            "message":        f"Username '{username}' is not registered. Please contact HR."
        }), 404

    if not row["email"]:
        conn.close()
        return jsonify({
            "success": False,
            "message": "No email address is registered for this account. Please contact HR directly."
        }), 400

    token  = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(32))
    expiry = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(
        "UPDATE users SET reset_token=?, reset_token_expiry=?, password_change_required=1 WHERE id=?",
        (token, expiry, row["id"])
    )
    conn.commit(); conn.close()

    reset_link = f"{PORTAL_URL}/?reset_token={token}&emp_id={row['id']}"
    ok, err    = send_email(row["email"], "MIIM — Password Reset Link",
                            reset_email_html(row["name"] or username, reset_link))
    if ok:
        return jsonify({"success": True,
                        "message": f"A reset link has been sent to {row['email']}. Please check your inbox."})
    else:
        return jsonify({"success": False,
                        "message": f"Could not send email: {err}. Please contact HR directly."}), 500


# ── RESET VIA TOKEN — validate token, set new temp password ──────────────────
@app.route("/reset_via_token", methods=["POST"])
def reset_via_token():
    """
    Body: { employee_id, token, new_password }
    Validates the token and sets the new password (still marked must-change).
    """
    data     = request.get_json() or {}
    emp_id   = data.get("employee_id")
    token    = data.get("token", "").strip()
    new_pw   = data.get("new_password", "")

    if not emp_id or not token or not new_pw:
        return jsonify({"success": False, "message": "employee_id, token and new_password required."}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, reset_token FROM users WHERE id = ? AND status = 'Active'", (emp_id,))
    row = cursor.fetchone()
    if not row or row["reset_token"] != token:
        conn.close()
        return jsonify({"success": False, "message": "Invalid or expired reset token."}), 401

    cursor.execute(
        "UPDATE users SET password = ?, password_change_required = 1, reset_token = NULL WHERE id = ?",
        (hash_password(new_pw), emp_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Temporary password set. Please log in and change your password."})


# ── SEND RESET LINK (HR triggers e-mail) ─────────────────────────────────────
@app.route("/send_reset_link", methods=["POST"])
def send_reset_link():
    """HR manually sends a password reset email from the dashboard.
    Body: { empId, token }"""
    import secrets, string
    data   = request.get_json() or {}
    emp_id = data.get("empId")
    token  = data.get("token", "").strip()
    if not emp_id:
        return jsonify({"success": False, "message": "empId required"}), 400

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT name, email, username, reset_token FROM users WHERE id = ?", (emp_id,))
    row = cursor.fetchone()

    if not row or not row["email"]:
        conn.close()
        return jsonify({"success": False, "message": "Employee or email not found."}), 404

    final_token = token if token else (row["reset_token"] or
        ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(32)))

    cursor.execute(
        "UPDATE users SET reset_token=?, password_change_required=1 WHERE id=?",
        (final_token, emp_id)
    )
    conn.commit(); conn.close()

    reset_link = f"{PORTAL_URL}/?reset_token={final_token}&emp_id={emp_id}"
    ok, err    = send_email(row["email"], "MIIM — HR Password Reset",
                            reset_email_html(row["name"] or row["username"], reset_link))
    if ok:
        return jsonify({"success": True,
                        "message": f"Reset link emailed to {row['email']}.",
                        "email": row["email"], "name": row["name"] or row["username"],
                        "reset_link": reset_link})
    else:
        return jsonify({"success": False,
                        "message": f"Email failed: {err}",
                        "reset_link": reset_link}), 500
@app.route("/upload_photo/<int:user_id>", methods=["POST"])
def upload_photo(user_id):
    if "photo" not in request.files:
        return jsonify({"success": False, "message": "No photo file provided."}), 400

    file = request.files["photo"]
    if not file.filename:
        return jsonify({"success": False, "message": "Empty filename."}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png"):
        return jsonify({"success": False, "message": "Only JPG/PNG allowed."}), 400

    save_ext = ".jpg" if ext in (".jpg", ".jpeg") else ".png"

    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return jsonify({"success": False, "message": "User not found."}), 404

    # Delete old photos for this user
    for fname in os.listdir(PHOTOS_DIR):
        if re.match(rf"^.+_{user_id}\.(jpg|png)$", fname):
            try: os.remove(os.path.join(PHOTOS_DIR, fname))
            except: pass

    safe_uname = re.sub(r"[^a-zA-Z0-9_]", "_", user["username"])
    filename   = f"{safe_uname}_{user_id}{save_ext}"
    file.save(os.path.join(PHOTOS_DIR, filename))

    cursor.execute("UPDATE users SET photo = ? WHERE id = ?", (filename, user_id))
    conn.commit()
    conn.close()
    build_excel()

    return jsonify({
        "success":   True,
        "message":   "Photo saved successfully.",
        "photo_url": f"/photo/{user_id}",
    })


@app.route("/photo/<int:user_id>")
def serve_photo(user_id):
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT photo FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()

    if not row or not row["photo"]:
        return jsonify({"error": "No photo"}), 404

    path = os.path.join(PHOTOS_DIR, row["photo"])
    if not os.path.exists(path):
        return jsonify({"error": "Photo file not found"}), 404

    mime = "image/png" if row["photo"].endswith(".png") else "image/jpeg"
    return send_file(path, mimetype=mime)


# =============================================================================
#  ROUTES — Excel
# =============================================================================
@app.route("/download_excel")
def download_excel():
    """Rebuild and download the latest Excel file."""
    ok, result = build_excel()
    if not ok:
        return jsonify({"error": f"Excel build failed: {result}"}), 500
    return send_file(
        EXCEL_FILE,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="MIIM_Employees.xlsx"
    )


@app.route("/sync_from_excel", methods=["POST"])
def sync_from_excel_route():
    """
    Manually trigger an Excel → DB sync.
    JSON response includes how many rows changed and a detailed log.
    """
    changes, log = sync_excel_to_db()
    return jsonify({
        "success": True,
        "changes": changes,
        "log":     log,
        "message": f"{changes} employee record(s) updated from Excel."
    })




# =============================================================================
#  EMERGENCY — Reset admin password back to admin/admin
#  GET http://localhost:5005/reset_admin
#  Only works from localhost. Remove in production.
# =============================================================================

# ── SET NEW PASSWORD (forced change — already authenticated) ──────────────────
@app.route("/set_new_password", methods=["POST"])
def set_new_password():
    data     = request.get_json() or {}
    username = data.get("username", "").strip()
    new_pw   = data.get("new_password", "")
    if not username or not new_pw:
        return jsonify({"success": False, "message": "username and new_password required."}), 400
    if len(new_pw) < 6:
        return jsonify({"success": False, "message": "Password must be at least 6 characters."}), 400
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM users WHERE username = ? AND status = 'Active'", (username,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "User not found."}), 404
    cursor.execute(
        "UPDATE users SET password = ?, password_change_required = 0, reset_token = NULL WHERE id = ?",
        (hash_password(new_pw), row["id"])
    )
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Password updated successfully."})

@app.route("/reset_admin")
def reset_admin():
    import socket
    if request.remote_addr not in ("127.0.0.1", "::1"):
        return jsonify({"error": "Forbidden"}), 403
    conn   = get_db()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET password=?, password_change_required=1 WHERE username='admin'",
        (hash_password("admin"),)
    )
    conn.commit(); conn.close()
    return jsonify({"success": True,
                    "message": "Admin password reset to 'admin'. You must set a new password on next login."})


# =============================================================================
#  ENTRY POINT
# =============================================================================
if __name__ == "__main__":
    init_db()
    print("\n🚀  MIIM HR Server  →  http://localhost:5005\n")
    app.run(debug=True, use_reloader=False, port=5005)